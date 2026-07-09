"""Excel 匯入:範本產生、逐列驗證、交易式入庫。

範本固定三列表頭:第 1 列欄名、第 2 列說明、第 3 列範例;
匯入時自第 4 列起讀取資料(前三列自動略過)。
驗證採「全對才寫入」:任一列有誤即回報所有錯誤,資料庫零寫入。
"""

import io
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.security import hash_password
from app.models.basedata import (
    ClassTrack,
    ClassUnit,
    RoomType,
    Subject,
    Teacher,
)
from app.models.period import PeriodTable
from app.models.user import Role, User, UserRole

HEADER_ROWS = 3  # 欄名 + 說明 + 範例

ROOM_TYPE_BY_LABEL = {
    "普通教室": RoomType.normal,
    "專科教室": RoomType.special,
    "實習工場": RoomType.workshop,
    "戶外": RoomType.outdoor,
}
TRACK_BY_LABEL = {
    "國小": ClassTrack.elementary,
    "國中": ClassTrack.junior_high,
    "普通型高中": ClassTrack.senior_high,
    "綜合型高中": ClassTrack.comprehensive,
    "技術型高中": ClassTrack.vocational,
}

# 每個實體的範本欄位:(欄名, 說明, 範例)
TEMPLATE_DEFS: dict[str, dict] = {
    "subjects": {
        "sheet": "科目",
        "columns": [
            ("名稱", "必填", "數學"),
            ("領域/群別", "選填", "數學領域"),
            ("需要場地類型", "選填:普通教室/專科教室/實習工場/戶外", "普通教室"),
            ("預設連堂", "選填,數字 1-8,預設 1", "1"),
        ],
    },
    "teachers": {
        "sheet": "教師",
        "columns": [
            ("姓名", "必填", "王小明"),
            ("身分末四碼", "選填,4 碼,用於辨識同名教師", "1234"),
            ("任教科目", "選填,多科以、分隔;需為已建立的科目", "數學、物理"),
            ("基本鐘點", "選填,數字", "20"),
            ("行政職稱", "選填", "教學組長"),
            ("行政減課", "選填,數字", "4"),
            ("外聘", "選填:是/否,預設否", "否"),
            ("登入帳號", "選填,勾選建立帳號時使用", "wang001"),
        ],
    },
    "classes": {
        "sheet": "班級",
        "columns": [
            ("年級", "必填,數字 1-12", "1"),
            ("班名", "必填", "甲"),
            ("學制", "必填:國小/國中/普通型高中/綜合型高中/技術型高中", "技術型高中"),
            ("群科", "選填(技高填寫)", "機械科"),
            ("導師", "選填,需為已建立的教師姓名", "王小明"),
            ("人數", "選填,數字", "35"),
            ("節次表", "選填,需為已建立的節次表名稱;空白則用學期預設", "高中部節次表"),
        ],
    },
}


@dataclass
class ImportResult:
    imported: int = 0
    errors: list[str] = field(default_factory=list)


def build_template(entity: str) -> bytes:
    cfg = TEMPLATE_DEFS[entity]
    wb = Workbook()
    ws = wb.active
    ws.title = cfg["sheet"]
    ws.append([c[0] for c in cfg["columns"]])
    ws.append([c[1] for c in cfg["columns"]])
    ws.append([c[2] for c in cfg["columns"]])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row: tuple, i: int) -> str | None:
    if i >= len(row) or row[i] is None:
        return None
    text = str(row[i]).strip()
    return text or None


def _parse_int(value: str | None) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))  # 容忍 Excel 把數字讀成 20.0
    except ValueError as err:
        raise ValueError(f"「{value}」不是有效數字") from err


def _data_rows(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    for idx, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True), start=HEADER_ROWS + 1
    ):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        yield idx, row


# ── 各實體匯入 ────────────────────────
def _import_subjects(db: Session, semester_id: int, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    pending: list[Subject] = []
    for idx, row in _data_rows(file_bytes):
        name = _cell(row, 0)
        if not name:
            result.errors.append(f"第 {idx} 列:名稱必填")
            continue
        room_label = _cell(row, 2)
        room_type = None
        if room_label:
            if room_label not in ROOM_TYPE_BY_LABEL:
                result.errors.append(f"第 {idx} 列:場地類型「{room_label}」無效")
                continue
            room_type = ROOM_TYPE_BY_LABEL[room_label].value
        try:
            block = _parse_int(_cell(row, 3)) or 1
        except ValueError as e:
            result.errors.append(f"第 {idx} 列:預設連堂 {e}")
            continue
        pending.append(
            Subject(
                semester_id=semester_id, name=name, domain=_cell(row, 1),
                required_room_type=room_type, default_block_size=block,
            )
        )
    if result.errors:
        return result
    db.add_all(pending)
    db.commit()
    result.imported = len(pending)
    return result


def _import_classes(db: Session, semester_id: int, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    teachers = {
        t.name: t.id
        for t in db.scalars(select(Teacher).where(Teacher.semester_id == semester_id))
    }
    period_tables = {
        pt.name: pt.id
        for pt in db.scalars(
            select(PeriodTable).where(PeriodTable.semester_id == semester_id)
        )
    }
    pending: list[ClassUnit] = []
    for idx, row in _data_rows(file_bytes):
        try:
            grade = _parse_int(_cell(row, 0))
        except ValueError as e:
            result.errors.append(f"第 {idx} 列:年級 {e}")
            continue
        name = _cell(row, 1)
        track_label = _cell(row, 2)
        if not grade or not name or not track_label:
            result.errors.append(f"第 {idx} 列:年級、班名、學制皆必填")
            continue
        if track_label not in TRACK_BY_LABEL:
            result.errors.append(f"第 {idx} 列:學制「{track_label}」無效")
            continue
        homeroom_name = _cell(row, 4)
        homeroom_id = None
        if homeroom_name:
            if homeroom_name not in teachers:
                result.errors.append(f"第 {idx} 列:導師「{homeroom_name}」不存在")
                continue
            homeroom_id = teachers[homeroom_name]
        table_name = _cell(row, 6)
        table_id = None
        if table_name:
            if table_name not in period_tables:
                result.errors.append(f"第 {idx} 列:節次表「{table_name}」不存在")
                continue
            table_id = period_tables[table_name]
        try:
            count = _parse_int(_cell(row, 5))
        except ValueError as e:
            result.errors.append(f"第 {idx} 列:人數 {e}")
            continue
        pending.append(
            ClassUnit(
                semester_id=semester_id, grade=grade, name=name,
                track=TRACK_BY_LABEL[track_label].value, department=_cell(row, 3),
                student_count=count, homeroom_teacher_id=homeroom_id,
                period_table_id=table_id,
            )
        )
    if result.errors:
        return result
    db.add_all(pending)
    db.commit()
    result.imported = len(pending)
    return result


def _import_teachers(
    db: Session, semester_id: int, file_bytes: bytes, create_accounts: bool
) -> ImportResult:
    result = ImportResult()
    subjects = {
        s.name: s
        for s in db.scalars(select(Subject).where(Subject.semester_id == semester_id))
    }
    existing_keys = {
        (t.name, t.id_last4 or "")
        for t in db.scalars(select(Teacher).where(Teacher.semester_id == semester_id))
    }
    existing_usernames = set(db.scalars(select(User.username)))

    seen_keys: set[tuple[str, str]] = set()
    seen_usernames: set[str] = set()
    pending: list[tuple[Teacher, str | None]] = []  # (teacher, username or None)

    for idx, row in _data_rows(file_bytes):
        name = _cell(row, 0)
        if not name:
            result.errors.append(f"第 {idx} 列:姓名必填")
            continue
        id_last4 = _cell(row, 1)
        key = (name, id_last4 or "")
        if key in existing_keys or key in seen_keys:
            result.errors.append(f"第 {idx} 列:教師「{name}」(末四碼 {id_last4 or '無'})重複")
            continue
        seen_keys.add(key)

        subject_objs: list[Subject] = []
        subj_field = _cell(row, 2)
        subj_error = False
        if subj_field:
            names = [s.strip() for s in subj_field.replace(",", "、").split("、") if s.strip()]
            for sname in names:
                if sname not in subjects:
                    result.errors.append(f"第 {idx} 列:科目「{sname}」不存在")
                    subj_error = True
                    break
                subject_objs.append(subjects[sname])
        if subj_error:
            continue
        try:
            base_periods = _parse_int(_cell(row, 3)) or 0
            admin_reduction = _parse_int(_cell(row, 5)) or 0
        except ValueError as e:
            result.errors.append(f"第 {idx} 列:{e}")
            continue
        is_external = (_cell(row, 6) or "否") == "是"

        username = _cell(row, 7)
        if create_accounts and username:
            if username in existing_usernames or username in seen_usernames:
                result.errors.append(f"第 {idx} 列:登入帳號「{username}」重複")
                continue
            seen_usernames.add(username)

        teacher = Teacher(
            semester_id=semester_id, name=name, id_last4=id_last4,
            base_periods=base_periods, admin_title=_cell(row, 4),
            admin_reduction=admin_reduction, is_external=is_external,
            subjects=subject_objs,
        )
        pending.append((teacher, username if create_accounts else None))

    if result.errors:
        return result

    for teacher, username in pending:
        db.add(teacher)
        if username:
            db.add(
                User(
                    username=username,
                    password_hash=hash_password(settings.default_import_password),
                    display_name=teacher.name,
                    must_change_password=True,
                    roles=[UserRole(role=Role.teacher.value)],
                )
            )
    db.commit()
    result.imported = len(pending)
    return result


def run_import(
    db: Session, entity: str, semester_id: int, file_bytes: bytes, create_accounts: bool = False
) -> ImportResult:
    if entity == "subjects":
        return _import_subjects(db, semester_id, file_bytes)
    if entity == "classes":
        return _import_classes(db, semester_id, file_bytes)
    if entity == "teachers":
        return _import_teachers(db, semester_id, file_bytes, create_accounts)
    raise ValueError(f"未知的匯入類型:{entity}")
