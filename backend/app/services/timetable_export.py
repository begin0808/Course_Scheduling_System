"""課表匯出:把已發布課表攤成格線,產生 Excel / PDF(HTML)/PNG(M5-1)。

三種對象(班級 / 教師 / 場地)共用同一個格線模型 `Grid`,再交給各格式的渲染器,
確保三種格式與畫面課表內容一致(驗收①)。資料來源一律是**已發布**課表(D4 快照)。

- Excel 由 openpyxl 產生,可在 api 同步跑(輕量)。
- PDF 需 WeasyPrint(系統依賴+中文字型只在 worker),故 PDF/PNG 走 worker 背景渲染;
  本模組只負責產出 HTML,實際 write_pdf 在 `app.workers.export_job`。
"""

import io
import zipfile
from dataclasses import dataclass, field
from datetime import date

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.basedata import ClassUnit, Room, Teacher
from app.models.period import Period, PeriodTable, PeriodType
from app.models.semester import Semester
from app.models.timetable import ScheduleEntry, Timetable, TimetableStatus
from app.services import period_tables as pt_service

_WEEKDAYS = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]


class ExportError(Exception):
    """匯出前置條件不成立(呼叫端轉為 4xx)。"""


# ── 格線模型 ────────────────────────────────────────────────
@dataclass(frozen=True, slots=True)
class Cell:
    lines: tuple[str, ...] = ()   # 科目 / 教師或班級 / 教室(逐行)
    span: int = 1                 # 連堂佔幾列
    covered: bool = False         # 被上方連堂覆蓋,不繪


@dataclass
class Row:
    period_no: int
    label: str
    is_regular: bool
    cells: list[Cell] = field(default_factory=list)  # 長度 = num_weekdays


@dataclass
class Grid:
    title: str
    num_weekdays: int
    rows: list[Row] = field(default_factory=list)

    @property
    def weekday_names(self) -> list[str]:
        return _WEEKDAYS[: self.num_weekdays]


@dataclass(frozen=True, slots=True)
class Meta:
    school_name: str
    semester_label: str
    timetable_name: str
    printed_on: date


@dataclass(frozen=True, slots=True)
class _EntryView:
    weekday: int
    period_no: int
    span: int
    subject: str
    teachers: str
    classes: str
    room: str
    teacher_ids: frozenset[int]
    class_ids: frozenset[int]
    room_id: int | None


class _Published:
    """一份已發布課表的匯出來源:格位、節次表、對象清單一次撈齊。"""

    def __init__(self, db: Session, semester_id: int) -> None:
        self.db = db
        semester = db.get(Semester, semester_id)
        if semester is None:
            raise ExportError("找不到學期")
        timetable = db.scalar(
            select(Timetable).where(
                Timetable.semester_id == semester_id,
                Timetable.status == TimetableStatus.published.value,
            )
        )
        if timetable is None:
            raise ExportError("此學期尚無已發布的課表")
        self.semester: Semester = semester
        self.timetable: Timetable = timetable
        self.entries = [
            _view(e)
            for e in db.scalars(
                select(ScheduleEntry).where(ScheduleEntry.timetable_id == self.timetable.id)
            )
        ]
        self.tables = {
            t.id: t for t in db.scalars(
                select(PeriodTable).where(PeriodTable.semester_id == semester_id)
            )
        }

    def meta(self) -> Meta:
        from app.core import clock
        return Meta(
            school_name=settings.school_name,
            semester_label=self.semester.label,
            timetable_name=self.timetable.name,
            printed_on=clock.school_today(),
        )

    def default_table(self) -> PeriodTable | None:
        tables = list(self.tables.values())
        return next((t for t in tables if t.is_default), tables[0] if tables else None)

    def class_table(self, cls: ClassUnit) -> PeriodTable | None:
        t = pt_service.resolve_period_table(self.db, cls)
        return t or self.default_table()


def _view(e: ScheduleEntry) -> _EntryView:
    a = e.assignment
    su = a.scheduling_unit
    room = e.room if e.room is not None else a.room
    return _EntryView(
        weekday=e.weekday, period_no=e.period_no, span=e.span,
        subject=a.subject.name,
        teachers="、".join(at.teacher.name for at in a.teachers),
        classes="、".join(m.class_unit.name for m in su.members),
        room=room.name if room else "",
        teacher_ids=frozenset(at.teacher_id for at in a.teachers),
        class_ids=frozenset(m.class_unit_id for m in su.members),
        room_id=e.effective_room_id,
    )


def _grid_from(
    table: PeriodTable | None,
    title: str,
    entries: list[_EntryView],
    lines_of,
) -> Grid:
    """把某對象的格位排進節次表格線。`lines_of(entry)` 決定每格顯示哪幾行。"""
    if table is None:
        return Grid(title=title, num_weekdays=5, rows=[])
    num_weekdays = table.num_weekdays
    periods = sorted(table.periods, key=lambda p: (p.period_no, p.weekday))
    # 每個 period_no 取一個代表(名稱/類型),weekday 小者優先
    by_no: dict[int, Period] = {}
    for p in periods:
        by_no.setdefault(p.period_no, p)
    order = sorted(by_no)

    # (weekday, period_no) → entry
    placed: dict[tuple[int, int], _EntryView] = {
        (e.weekday, e.period_no): e for e in entries
    }
    covered: set[tuple[int, int]] = set()
    for e in entries:
        for k in range(1, e.span):
            covered.add((e.weekday, e.period_no + k))

    grid = Grid(title=title, num_weekdays=num_weekdays)
    for pno in order:
        rep = by_no[pno]
        is_regular = rep.type == PeriodType.regular.value
        row = Row(period_no=pno, label=rep.name, is_regular=is_regular)
        for wd in range(1, num_weekdays + 1):
            if (wd, pno) in covered:
                row.cells.append(Cell(covered=True))
            elif (wd, pno) in placed:
                e = placed[(wd, pno)]
                row.cells.append(Cell(lines=tuple(lines_of(e)), span=e.span))
            else:
                row.cells.append(Cell())
        grid.rows.append(row)
    return grid


# ── 三種對象 → Grid ─────────────────────────────────────────
def _class_lines(e: _EntryView) -> list[str]:
    return [x for x in (e.subject, e.teachers, e.room) if x]


def _teacher_lines(e: _EntryView) -> list[str]:
    return [x for x in (e.subject, e.classes, e.room) if x]


def _room_lines(e: _EntryView) -> list[str]:
    return [x for x in (e.subject, e.classes, e.teachers) if x]


def class_grid(pub: _Published, cls: ClassUnit) -> Grid:
    entries = [e for e in pub.entries if cls.id in e.class_ids]
    return _grid_from(pub.class_table(cls), f"{cls.grade}年{cls.name} 課表", entries, _class_lines)


def teacher_grid(pub: _Published, teacher: Teacher) -> Grid:
    entries = [e for e in pub.entries if teacher.id in e.teacher_ids]
    return _grid_from(pub.default_table(), f"{teacher.name} 課表", entries, _teacher_lines)


def room_grid(pub: _Published, room: Room) -> Grid:
    entries = [e for e in pub.entries if e.room_id == room.id]
    return _grid_from(pub.default_table(), f"{room.name} 課表", entries, _room_lines)


def build_grid(db: Session, semester_id: int, view: str, target_id: int) -> tuple[Grid, Meta]:
    pub = _Published(db, semester_id)
    if view == "class":
        obj = db.get(ClassUnit, target_id)
        if obj is None or obj.semester_id != semester_id:
            raise ExportError("找不到班級")
        return class_grid(pub, obj), pub.meta()
    if view == "teacher":
        obj_t = db.get(Teacher, target_id)
        if obj_t is None or obj_t.semester_id != semester_id:
            raise ExportError("找不到教師")
        return teacher_grid(pub, obj_t), pub.meta()
    if view == "room":
        obj_r = db.get(Room, target_id)
        if obj_r is None or obj_r.semester_id != semester_id:
            raise ExportError("找不到場地")
        return room_grid(pub, obj_r), pub.meta()
    raise ExportError(f"未知的檢視類型:{view}")


def _classes(db: Session, semester_id: int) -> list[ClassUnit]:
    return list(db.scalars(
        select(ClassUnit).where(ClassUnit.semester_id == semester_id)
        .order_by(ClassUnit.grade, ClassUnit.name)
    ))


def school_workbook(db: Session, semester_id: int) -> bytes:
    """全校總表:一個 Excel 檔,每班一個分頁。"""
    pub = _Published(db, semester_id)
    grids = [class_grid(pub, c) for c in _classes(db, semester_id)]
    return grids_to_xlsx(grids, pub.meta())


def class_batch_zip(db: Session, semester_id: int) -> bytes:
    """批次匯出:全部班級各一個 Excel 檔,打包成 zip。"""
    pub = _Published(db, semester_id)
    meta = pub.meta()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for c in _classes(db, semester_id):
            grid = class_grid(pub, c)
            z.writestr(f"{c.grade}年{c.name}.xlsx", grids_to_xlsx([grid], meta))
    return buf.getvalue()


# ── Excel 渲染 ──────────────────────────────────────────────
def _safe_sheet_title(title: str, used: set[str]) -> str:
    # Excel 分頁名 ≤31 字、不可含 : \ / ? * [ ]
    clean = title
    for ch in ':\\/?*[]':
        clean = clean.replace(ch, " ")
    clean = clean[:28].strip() or "課表"
    name, i = clean, 1
    while name in used:
        name = f"{clean[:26]}~{i}"
        i += 1
    used.add(name)
    return name


def grids_to_xlsx(grids: list[Grid], meta: Meta) -> bytes:
    """每張 Grid 一個分頁。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    head_fill = PatternFill("solid", fgColor="E8E8E8")
    shade = PatternFill("solid", fgColor="F5F5F5")
    used: set[str] = set()

    for grid in grids:
        ws = wb.create_sheet(_safe_sheet_title(grid.title, used))
        ws.append([grid.title])
        ws.append([f"{meta.school_name}　{meta.semester_label}　列印日:{meta.printed_on}"])
        ws["A1"].font = Font(bold=True, size=14)
        header = ["節次", *grid.weekday_names]
        ws.append(header)
        head_row = ws.max_row
        for c in ws[head_row]:
            c.font = Font(bold=True)
            c.fill = head_fill
            c.alignment = center
            c.border = border

        for row in grid.rows:
            values = [row.label]
            for cell in row.cells:
                values.append("\n".join(cell.lines) if cell.lines else "")
            ws.append(values)
            r = ws.max_row
            for ci, c in enumerate(ws[r]):
                c.alignment = center
                c.border = border
                if ci == 0 or not row.is_regular:
                    c.fill = shade
            # 連堂:垂直合併
            for ci, cell in enumerate(row.cells):
                if cell.span > 1:
                    col = ci + 2
                    ws.merge_cells(start_row=r, start_column=col,
                                   end_row=r + cell.span - 1, end_column=col)

        ws.column_dimensions["A"].width = 10
        for col in range(2, 2 + grid.num_weekdays):
            ws.column_dimensions[ws.cell(row=head_row, column=col).column_letter].width = 16
        ws.freeze_panes = ws.cell(row=head_row + 1, column=2)

    if not wb.sheetnames:  # 全空:給一張空白頁避免壞檔
        wb.create_sheet("課表")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTML 渲染(供 worker 轉 PDF/PNG)─────────────────────────
def _esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def grid_to_html(grid: Grid, meta: Meta) -> str:
    """單一對象的 A4 直式課表 HTML。中文由 worker 映像內嵌的 Noto CJK 呈現。"""
    head_cells = "".join(f"<th>{_esc(w)}</th>" for w in grid.weekday_names)
    subtitle = (f"{_esc(meta.school_name)}　{_esc(meta.semester_label)}　"
                f"{_esc(meta.timetable_name)}　列印日:{meta.printed_on}")
    body_rows = []
    for row in grid.rows:
        tds = [f'<th class="pno">{_esc(row.label)}</th>']
        for cell in row.cells:
            if cell.covered:
                continue
            cls = "cell" if row.is_regular else "cell rest"
            span = f' rowspan="{cell.span}"' if cell.span > 1 else ""
            inner = "<br>".join(_esc(x) for x in cell.lines)
            tds.append(f'<td class="{cls}"{span}>{inner}</td>')
        body_rows.append(f"<tr>{''.join(tds)}</tr>")
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>
@page {{ size: A4 portrait; margin: 14mm; }}
* {{ font-family: "Noto Sans CJK TC", "Noto Sans TC", sans-serif; }}
h1 {{ font-size: 18px; text-align: center; margin: 0 0 4px; }}
.meta {{ text-align: center; font-size: 12px; color: #444; margin-bottom: 10px; }}
table {{ border-collapse: collapse; width: 100%; table-layout: fixed; }}
th, td {{ border: 1px solid #333; padding: 4px; text-align: center; font-size: 12px;
  vertical-align: middle; word-break: break-all; }}
thead th {{ background: #e8e8e8; }}
.pno {{ background: #f2f2f2; width: 60px; font-size: 11px; }}
td.cell {{ height: 46px; }}
td.rest {{ background: #f7f7f7; color: #888; }}
</style></head><body>
<h1>{_esc(grid.title)}</h1>
<div class="meta">{subtitle}</div>
<table><thead><tr><th class="pno">節次</th>{head_cells}</tr></thead>
<tbody>{''.join(body_rows)}</tbody></table>
</body></html>"""
