"""代課鐘點月結統計(M4-5,architecture.md §5.4)。

回答「這個月每位老師代了幾節、其中幾節要計鐘點費」。真相仍是 `substitution` 列
(處置決定)+ `affected_period`(受影響節次快照),這裡依教師彙總。

**兩個數字**:
- 代課節數:該教師接手的所有處置(代課/調課/併班),即他實際處理了幾節。
- 計費節數:其中 `counts_toward_hours` 為真者。併班/自習預設不計、代課預設計(可覆寫)。

**跨月假單自動拆月**:以每一個 `affected_period` 自己的日期分月,不是以假單分月。
王師請 1/30~2/2 的假,1 月的節次進 1 月報表、2 月的進 2 月——不必特別處理。

**銷假的節次不計**:銷假會把未完成的節次轉為 `cancelled`(那堂課沒上),故排除;
已完成(completed)的節次即使事後銷假仍保留(課上過了,鐘點照算)。
"""

import io
from dataclasses import dataclass, field
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.leave import LEAVE_TYPE_CN, AffectedPeriod, AffectedStatus
from app.models.substitution import SUBSTITUTION_TYPE_CN, Substitution

_Date = date


@dataclass(frozen=True, slots=True)
class StatDetail:
    """一列代課明細:某教師某節的處置。"""

    handler_teacher_id: int
    handler_name: str
    date: _Date
    period_no: int
    period_name: str
    class_names: str
    subject_name: str
    absent_teacher_name: str
    leave_type: str
    leave_type_label: str
    sub_type: str
    sub_type_label: str
    counts_toward_hours: bool
    funding_source: str


@dataclass
class TeacherSummary:
    teacher_id: int
    teacher_name: str
    handled_count: int = 0   # 代課節數(所有接手處置)
    billable_count: int = 0  # 計費節數(counts_toward_hours 為真)


@dataclass
class MonthlyReport:
    year: int
    month: int
    summaries: list[TeacherSummary] = field(default_factory=list)
    details: list[StatDetail] = field(default_factory=list)


def _next_month(year: int, month: int) -> date:
    return date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)


def monthly_report(
    db: Session,
    semester_id: int,
    year: int,
    month: int,
    *,
    teacher_id: int | None = None,
) -> MonthlyReport:
    """某月的代課鐘點統計。指定 teacher_id 則只統計該教師(教師個人查詢用)。"""
    month_start = date(year, month, 1)
    month_end = _next_month(year, month)

    stmt = (
        select(Substitution, AffectedPeriod)
        .join(AffectedPeriod, Substitution.affected_period_id == AffectedPeriod.id)
        .where(
            Substitution.semester_id == semester_id,
            Substitution.handler_teacher_id.isnot(None),
            AffectedPeriod.status != AffectedStatus.cancelled.value,
            AffectedPeriod.date >= month_start,
            AffectedPeriod.date < month_end,
        )
    )
    if teacher_id is not None:
        stmt = stmt.where(Substitution.handler_teacher_id == teacher_id)

    report = MonthlyReport(year=year, month=month)
    summaries: dict[int, TeacherSummary] = {}

    for sub, ap in db.execute(stmt).all():
        handler = sub.handler
        if handler is None:  # handler 已被移除(SET NULL 尚未反映在關聯)
            continue
        leave = ap.leave_request
        report.details.append(StatDetail(
            handler_teacher_id=handler.id,
            handler_name=handler.name,
            date=ap.date,
            period_no=ap.period_no,
            period_name=ap.period_name,
            class_names=ap.class_names,
            subject_name=ap.subject_name,
            absent_teacher_name=leave.teacher.name if leave.teacher else "(已移除)",
            leave_type=leave.leave_type,
            leave_type_label=LEAVE_TYPE_CN.get(leave.leave_type, leave.leave_type),
            sub_type=sub.type,
            sub_type_label=SUBSTITUTION_TYPE_CN.get(sub.type, sub.type),
            counts_toward_hours=sub.counts_toward_hours,
            funding_source=sub.funding_source,
        ))
        s = summaries.get(handler.id)
        if s is None:
            s = TeacherSummary(teacher_id=handler.id, teacher_name=handler.name)
            summaries[handler.id] = s
        s.handled_count += 1
        if sub.counts_toward_hours:
            s.billable_count += 1

    report.details.sort(key=lambda d: (d.handler_name, d.date, d.period_no))
    report.summaries = sorted(summaries.values(), key=lambda s: s.teacher_name)
    return report


_DETAIL_HEADERS = (
    "教師", "日期", "節次", "班級", "科目", "原任教師", "假別", "處置", "計費", "經費來源",
)
_SUMMARY_HEADERS = ("教師", "代課節數", "計費節數")


def build_workbook(report: MonthlyReport) -> bytes:
    """匯出兩張表:彙總(每位教師)+ 明細(逐節)。回傳 xlsx bytes。"""
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "彙總"
    ws_sum.append(list(_SUMMARY_HEADERS))
    for s in report.summaries:
        ws_sum.append([s.teacher_name, s.handled_count, s.billable_count])

    ws_detail = wb.create_sheet("明細")
    ws_detail.append(list(_DETAIL_HEADERS))
    for d in report.details:
        ws_detail.append([
            d.handler_name, d.date.isoformat(), d.period_name, d.class_names, d.subject_name,
            d.absent_teacher_name, d.leave_type_label, d.sub_type_label,
            "是" if d.counts_toward_hours else "否", d.funding_source,
        ])

    for ws in (ws_sum, ws_detail):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
