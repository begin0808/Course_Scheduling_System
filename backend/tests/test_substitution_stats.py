"""M4-5:代課鐘點月結統計。

重點:兩個數字(代課節數 vs 計費節數)、不計費項排除(併班/自習)、跨月假單拆月計、
教師個人只看自己、Excel 匯出欄位。
"""

import io

import pytest
from openpyxl import load_workbook

from app.models.basedata import Teacher
from app.models.user import Role
from tests.conftest import make_user

# 日期一律由執行當日推算,不硬編(見 tests/dates.py)。
# CROSS_WED / CROSS_WED2 是相鄰但分屬不同月份的兩個週三,供跨月拆帳驗證。
from tests.dates import CROSS_WED, CROSS_WED2, SEM_END, SEM_START, WED
from tests.test_substitutions import _World

PW = "password123"


def _other_month(day):
    """一個確定沒有任何代課紀錄的月份(基準日的下個月)。"""
    return (day.year + 1, 1) if day.month == 12 else (day.year, day.month + 1)


@pytest.fixture
def w(env):
    client, db = env
    make_user(db, "s", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "s", "password": PW})
    sid = client.post("/api/semesters", json={
        "academic_year": 115, "term": 1, "template_key": "junior_high",
        "start_date": SEM_START.isoformat(), "end_date": SEM_END.isoformat(),
    }).json()["id"]
    return _World(client, db, sid)


def _stats(w, year=None, month=None, **params):
    """預設查「基準週那個月」——受影響節次都落在那裡。"""
    year = WED.year if year is None else year
    month = WED.month if month is None else month
    qs = "".join(f"&{k}={v}" for k, v in params.items() if v is not None)
    return w.client.get(f"/api/substitution-stats{w.q}&year={year}&month={month}{qs}").json()


# ── 驗收①:代課節數 vs 計費節數,不計費項排除 ──────────────
def test_handled_vs_billable_counts(w):
    """陳師代課 1 節(計費)+ 併班 1 節(不計費);自習無處理教師不計入任何人。"""
    w.teacher("王師", ["國文"])
    w.teacher("陳師", ["國文"])
    w.place("王師", "國文", "701", 0)  # 週三第一節
    w.place("王師", "國文", "702", 1)  # 週三第二節
    w.place("王師", "國文", "703", 2)  # 週三第三節
    w.publish()
    ap = w.leave("王師")  # 3 節受影響(同一天)

    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陳師"])
    w.assign(ap[1]["id"], type="merge", handler_teacher_id=w.teachers["陳師"])
    w.assign(ap[2]["id"], type="self_study")  # 無處理教師

    data = _stats(w)
    chen = next(s for s in data["summaries"] if s["teacher_name"] == "陳師")
    assert chen["handled_count"] == 2   # 代課 + 併班
    assert chen["billable_count"] == 1  # 只有代課計費
    # 自習沒有處理教師,不會出現在任何人的統計
    assert all(s["teacher_name"] != "王師" for s in data["summaries"])
    assert len(data["details"]) == 2


def test_detail_columns(w):
    w.teacher("王師", ["國文"])
    w.teacher("陳師", ["國文"])
    w.place("王師", "國文", "701", 0)
    w.publish()
    ap = w.leave("王師")
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陳師"])

    d = _stats(w)["details"][0]
    assert d["handler_name"] == "陳師"
    assert d["absent_teacher_name"] == "王師"
    assert d["leave_type_label"] == "病假"
    assert d["sub_type_label"] == "代課"
    assert d["counts_toward_hours"] is True
    assert d["class_names"] == "701"
    assert d["subject_name"] == "國文"


def test_month_filter_excludes_other_months(w):
    w.teacher("王師", ["國文"])
    w.teacher("陳師", ["國文"])
    w.place("王師", "國文", "701", 0)
    w.publish()
    ap = w.leave("王師", when=WED)
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陳師"])

    other_year, other_month = _other_month(WED)
    assert _stats(w)["summaries"]                                          # 當月有
    assert _stats(w, year=other_year, month=other_month)["summaries"] == []  # 別的月份沒有


# ── 邊界:跨月假單拆月計 ─────────────────────────────────────
def test_cross_month_leave_splits_by_period_date(w):
    """林師請的假橫跨月底(含兩個週三,分屬前後兩個月),各月各計一節。

    分月依據是「每個受影響節次自己的日期」,不是假單的起始月。
    """
    assert CROSS_WED.month != CROSS_WED2.month, "前置條件:兩個週三必須跨月"
    w.teacher("林師", ["數學"])
    w.teacher("周師", ["數學"])
    w.place("林師", "數學", "701", 0, weekday=3)  # 每週三第一節
    w.publish()

    leave = w.client.post(f"/api/leaves{w.q}", json={
        "teacher_id": w.teachers["林師"], "leave_type": "official",
        "start_date": CROSS_WED.isoformat(), "end_date": CROSS_WED2.isoformat()}).json()
    aps = leave["affected_periods"]
    dates = sorted(p["date"] for p in aps)
    assert dates == [CROSS_WED.isoformat(), CROSS_WED2.isoformat()], dates
    for p in aps:
        w.assign(p["id"], type="substitute", handler_teacher_id=w.teachers["周師"])

    first = _stats(w, year=CROSS_WED.year, month=CROSS_WED.month)["summaries"]
    second = _stats(w, year=CROSS_WED2.year, month=CROSS_WED2.month)["summaries"]
    assert next(s for s in first if s["teacher_name"] == "周師")["billable_count"] == 1
    assert next(s for s in second if s["teacher_name"] == "周師")["billable_count"] == 1


# ── 銷假的節次不計 ───────────────────────────────────────────
def test_cancelled_leave_excluded(w):
    w.teacher("王師", ["國文"])
    w.teacher("陳師", ["國文"])
    w.place("王師", "國文", "701", 0)
    w.publish()
    leave = w.client.post(f"/api/leaves{w.q}", json={
        "teacher_id": w.teachers["王師"], "leave_type": "sick",
        "start_date": WED.isoformat(), "end_date": WED.isoformat()}).json()
    w.assign(leave["affected_periods"][0]["id"],
             type="substitute", handler_teacher_id=w.teachers["陳師"])
    assert _stats(w)["summaries"], "銷假前應計入"

    w.client.post(f"/api/leaves/{leave['id']}/cancel")
    assert _stats(w)["summaries"] == [], "銷假後該節不計(那堂課沒上)"


# ── 教師個人查詢:只看自己 ──────────────────────────────────
def test_mine_shows_only_own(w):
    w.teacher("王師", ["國文"])
    w.teacher("陳師", ["國文"])
    w.teacher("周師", ["國文"])
    w.place("王師", "國文", "701", 0)
    w.place("王師", "國文", "702", 1)
    w.publish()
    ap = w.leave("王師")
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陳師"])
    w.assign(ap[1]["id"], type="substitute", handler_teacher_id=w.teachers["周師"])

    # 綁定 chenacc 帳號到陳師
    user = make_user(w.db, "chenacc", PW, roles=[Role.teacher])
    chen = w.db.get(Teacher, w.teachers["陳師"])
    chen.user_id = user.id
    w.db.commit()

    w.client.post("/api/auth/logout")
    w.client.post("/api/auth/login", json={"username": "chenacc", "password": PW})
    mine = w.client.get(
        f"/api/substitution-stats/mine{w.q}&year={WED.year}&month={WED.month}").json()
    names = {d["handler_name"] for d in mine["details"]}
    assert names == {"陳師"}, names
    assert len(mine["details"]) == 1


def test_teacher_cannot_view_full_stats(w):
    make_user(w.db, "t", PW, roles=[Role.teacher])
    w.client.post("/api/auth/logout")
    w.client.post("/api/auth/login", json={"username": "t", "password": PW})
    r = w.client.get(f"/api/substitution-stats{w.q}&year={WED.year}&month={WED.month}")
    assert r.status_code == 403


# ── Excel 匯出 ───────────────────────────────────────────────
def test_export_xlsx(w):
    w.teacher("王師", ["國文"])
    w.teacher("陳師", ["國文"])
    w.place("王師", "國文", "701", 0)
    w.publish()
    ap = w.leave("王師")
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陳師"])

    r = w.client.get(
        f"/api/substitution-stats/export{w.q}&year={WED.year}&month={WED.month}")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["彙總", "明細"]
    detail = wb["明細"]
    headers = [c.value for c in detail[1]]
    assert headers[:8] == ["教師", "日期", "節次", "班級", "科目", "原任教師", "假別", "處置"]
    assert "計費" in headers
    row = [c.value for c in detail[2]]
    assert row[0] == "陳師"
    assert row[5] == "王師"       # 原任教師
    assert row[8] == "是"         # 計費
