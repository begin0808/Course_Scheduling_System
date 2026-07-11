"""M5-1:課表匯出。Excel 以讀回比對驗證內容;HTML(PDF 來源)驗中文與版面;
全校總表/批次/RBAC。PDF/PNG 的 worker 渲染於真實環境另行驗證。
"""

import io
import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook

from app.models.user import Role
from app.services import timetable_export as tex
from tests.conftest import make_user
from tests.test_substitutions import _World

PW = "password123"
SEM_START = date(2026, 9, 1)
SEM_END = date(2027, 1, 20)


@pytest.fixture
def w(env):
    client, db = env
    make_user(db, "s", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "s", "password": PW})
    sid = client.post("/api/semesters", json={
        "academic_year": 115, "term": 1, "template_key": "junior_high",
        "start_date": SEM_START.isoformat(), "end_date": SEM_END.isoformat(),
    }).json()["id"]
    world = _World(client, db, sid)
    world.teacher("王師", ["國文"])
    world.place("王師", "國文", "701", 0)   # 週三第一節
    world.publish()
    return world


def _cells(ws) -> set[str]:
    out: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v:
                out.add(str(v))
    return out


def _xlsx(resp):
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    return load_workbook(io.BytesIO(resp.content))


# ── 內容一致(Excel 讀回)─────────────────────────────────────
def test_class_excel_matches_grid(w):
    cid = w.classes["701"]
    r = w.client.get(f"/api/export/timetable{w.q}&view=class&target_id={cid}&fmt=xlsx")
    wb = _xlsx(r)
    ws = wb.active
    cells = _cells(ws)
    assert "701 課表" in " ".join(cells) or any("課表" in c for c in cells)
    # 週三第一節有「國文」與「王師」
    joined = "\n".join(cells)
    assert "國文" in joined
    assert "王師" in joined
    assert "週三" in cells  # 表頭星期
    assert "第一節" in cells  # 節次名


def test_teacher_excel_shows_class_not_teacher(w):
    tid = w.teachers["王師"]
    r = w.client.get(f"/api/export/timetable{w.q}&view=teacher&target_id={tid}&fmt=xlsx")
    wb = _xlsx(r)
    joined = "\n".join(_cells(wb.active))
    assert "國文" in joined
    assert "701" in joined       # 教師視角格內顯示班級


def test_room_export_ok_even_without_room(w):
    # 建一個沒有課的場地也應匯出成功(空課表)
    rid = w.client.post(f"/api/rooms{w.q}", json={"name": "101教室"}).json()["id"]
    r = w.client.get(f"/api/export/timetable{w.q}&view=room&target_id={rid}&fmt=xlsx")
    assert r.status_code == 200


# ── HTML(PDF 來源)─────────────────────────────────────────
def test_grid_html_has_chinese_and_layout(w):
    cid = w.classes["701"]
    grid, meta = tex.build_grid(w.db, w.sid, "class", cid)
    html = tex.grid_to_html(grid, meta)
    assert "國文" in html and "王師" in html
    assert "週三" in html and "第一節" in html
    assert "A4 portrait" in html
    assert meta.school_name in html


# ── 全校總表 / 批次 ─────────────────────────────────────────
def test_school_workbook_one_sheet_per_class(w):
    # 學期已有 701 與佔位班 900 → 全校總表每班一分頁
    r = w.client.get(f"/api/export/school.xlsx{w.q}")
    wb = _xlsx(r)
    assert len(wb.sheetnames) >= 2
    assert any("701" in s for s in wb.sheetnames)


def test_batch_zip_has_per_class_files(w):
    r = w.client.get(f"/api/export/batch.zip{w.q}")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"
    z = zipfile.ZipFile(io.BytesIO(r.content))
    names = z.namelist()
    assert any(n.endswith(".xlsx") for n in names)
    assert any("701" in n for n in names)


# ── 邊界 / RBAC ─────────────────────────────────────────────
def test_export_unpublished_semester_404(env):
    client, db = env
    make_user(db, "s2", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "s2", "password": PW})
    sid = client.post("/api/semesters", json={
        "academic_year": 116, "term": 1, "template_key": "junior_high",
        "start_date": SEM_START.isoformat(), "end_date": SEM_END.isoformat(),
    }).json()["id"]
    cid = client.post(f"/api/class-units?semester_id={sid}",
                      json={"grade": 7, "name": "701", "track": "junior_high"}).json()["id"]
    r = client.get(f"/api/export/timetable?semester_id={sid}&view=class&target_id={cid}&fmt=xlsx")
    assert r.status_code == 404


def test_batch_60_classes_under_60s(db):
    """驗收②:60 班批次匯出 < 60 秒(Excel 產生為 CPU-bound,與資料庫無關)。"""
    import time

    from app.models.timetable import Timetable, TimetableStatus
    from tests.fixtures import build_large_school

    fx = build_large_school(db, num_classes=60)
    db.add(Timetable(
        semester_id=fx.semester_id, name="正式課表",
        status=TimetableStatus.published.value))
    db.commit()

    t0 = time.perf_counter()
    data = tex.class_batch_zip(db, fx.semester_id)
    elapsed = time.perf_counter() - t0
    assert elapsed < 60, f"批次匯出耗時 {elapsed:.1f}s"
    z = zipfile.ZipFile(io.BytesIO(data))
    assert len(z.namelist()) == 60


def test_teacher_can_export_single_but_not_batch(w):
    make_user(w.db, "t", PW, roles=[Role.teacher])
    w.client.post("/api/auth/logout")
    w.client.post("/api/auth/login", json={"username": "t", "password": PW})
    cid = w.classes["701"]
    single = w.client.get(f"/api/export/timetable{w.q}&view=class&target_id={cid}&fmt=xlsx")
    assert single.status_code == 200          # 單一課表全校可查可匯出
    assert w.client.get(f"/api/export/school.xlsx{w.q}").status_code == 403
    assert w.client.get(f"/api/export/batch.zip{w.q}").status_code == 403
